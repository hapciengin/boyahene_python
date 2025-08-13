import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
import os
import sys

# ==============================================================================
# --- YAPILANDIRMA BÖLÜMÜ ---
# ==============================================================================
DEFAULT_NETWORK_PATH = r"\\10.0.1.157\Test"
EXCEL_SHEETNAME = "Barkod Verileri"
COLUMN_HEADERS = [
    "Parça 1", "Parça 2", "Parça 3", "Parça 4", "Parça 5",
    "Okutulan Ham Barkod"
]
BARCODE_SLICING_RULES = {
    "Parça 1": (0, 5), "Parça 2": (5, 10), "Parça 3": (10, 15),
    "Parça 4": (15, 20), "Parça 5": (-7, None),
}
PART_TRANSLATIONS = {
    "Parça 1": {"12345": "Kırmızı Renk", "67890": "Mavi Renk"},
    "Parça 2": {"ABCDE": "İnce Kalınlık", "FGHIJ": "Standart Kalınlık"},
}
# ==============================================================================
# --- PROGRAM KODLARI ---
# ==============================================================================
def extract_and_slice_barcode(barcode_str: str) -> dict:
    digits_only = "".join(filter(str.isdigit, barcode_str))
    sliced_data = {}
    for header in COLUMN_HEADERS:
        if header == "Okutulan Ham Barkod": sliced_data[header] = barcode_str; continue
        sliced_part = ""
        rule = BARCODE_SLICING_RULES.get(header)
        if rule: start, end = rule; sliced_part = digits_only[start:end]
        translated_part = PART_TRANSLATIONS.get(header, {}).get(sliced_part, sliced_part)
        sliced_data[header] = translated_part
    return sliced_data

def save_to_excel(data_dict: dict, excel_filepath: str):
    try:
        if not os.path.exists(excel_filepath):
            workbook = openpyxl.Workbook(); sheet = workbook.active; sheet.title = EXCEL_SHEETNAME; sheet.append(COLUMN_HEADERS)
        else:
            workbook = openpyxl.load_workbook(excel_filepath)
            sheet = workbook[EXCEL_SHEETNAME] if EXCEL_SHEETNAME in workbook.sheetnames else workbook.create_sheet(EXCEL_SHEETNAME)
            if sheet.max_row == 0: sheet.append(COLUMN_HEADERS)
        sheet.append([data_dict.get(header, "") for header in COLUMN_HEADERS])
        workbook.save(excel_filepath)
        return "eklendi"
    except PermissionError: return "izin_hatasi"
    except Exception: return "hata"

def delete_from_excel(barcode_to_delete: str, excel_filepath: str):
    if not os.path.exists(excel_filepath): return "dosya_yok"
    try:
        workbook = openpyxl.load_workbook(excel_filepath); sheet = workbook[EXCEL_SHEETNAME]
        header_row = sheet[1]; barcode_col_idx = -1
        for idx, cell in enumerate(header_row):
            if cell.value == "Okutulan Ham Barkod": barcode_col_idx = idx + 1; break
        if barcode_col_idx == -1: return "baslik_yok"
        row_to_delete_idx = -1
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=barcode_col_idx).value == barcode_to_delete: row_to_delete_idx = i; break
        if row_to_delete_idx != -1: sheet.delete_rows(row_to_delete_idx); workbook.save(excel_filepath); return "silindi"
        else: return "bulunamadi"
    except PermissionError: return "izin_hatasi"
    except Exception: return "hata"

class BarcodeApp(tk.Tk):
    def __init__(self):
        super().__init__(); self.excel_filepath = None; self.current_mode = "EKLEME"
        self.title("Boya Barkod Yönetimi (v8.0 - Ağ/Test Klasörü)"); self.geometry("650x350"); self.configure(bg='#f0f0f0')
        main_frame = tk.Frame(self, padx=15, pady=15, bg='#f0f0f0'); main_frame.pack(expand=True, fill=tk.BOTH)
        
        network_frame = tk.Frame(main_frame, bg='#f0f0f0'); network_frame.pack(fill='x', pady=(0, 5))
        tk.Label(network_frame, text="Ağ Yolu:", font=("Helvetica", 10, "bold"), bg='#f0f0f0').pack(side='left')
        self.network_path_var = tk.StringVar(value=DEFAULT_NETWORK_PATH)
        self.network_path_entry = tk.Entry(network_frame, textvariable=self.network_path_var, font=("Helvetica", 9)); self.network_path_entry.pack(side='left', fill='x', expand=True, padx=5)
        scan_button = tk.Button(network_frame, text="Bu Yolu Tara", command=self.scan_manual_path, font=("Helvetica", 9, "bold")); scan_button.pack(side='left')

        file_frame = tk.Frame(main_frame, bg='#f0f0f0'); file_frame.pack(fill='x', pady=5)
        select_button = tk.Button(file_frame, text="Veya Dosyayı Manuel Seç / Oluştur...", command=self.select_file, font=("Helvetica", 10, "bold"), bg="#5cb85c", fg="white"); select_button.pack(side='left', padx=(0, 10))
        self.file_label = tk.Label(file_frame, text="...", font=("Helvetica", 9), bg='white', relief='sunken', anchor='w', padx=5); self.file_label.pack(side='left', fill='x', expand=True)

        self.mode_label = tk.Label(main_frame, text="MOD: EKLEME (2)", font=("Helvetica", 14, "bold"), fg="green", bg='#f0f0f0'); self.mode_label.pack(pady=10)
        tk.Label(main_frame, text="BARKODU OKUTUN", font=("Helvetica", 12, "bold"), bg='#f0f0f0').pack()
        self.barcode_entry = tk.Entry(main_frame, font=("Arial", 16), bd=2, relief=tk.GROOVE); self.barcode_entry.pack(padx=5, pady=5, fill=tk.X)
        self.status_label = tk.Label(main_frame, text="Başlatılıyor...", font=("Helvetica", 11), fg="navy", bg='#f0f0f0', height=2); self.status_label.pack(pady=10)
        
        self.bind_all('1', self.set_delete_mode); self.bind_all('2', self.set_add_mode); self.bind('<Return>', self.process_barcode)
        self.barcode_entry.focus_set()
        self.after(200, self.scan_manual_path)

    def set_active_file(self, filepath):
        self.excel_filepath = filepath
        self.file_label.config(text=f" Aktif Dosya: {filepath}", fg='green')
        self.set_add_mode()

    def scan_manual_path(self):
        path_to_scan = self.network_path_var.get().strip()
        if not path_to_scan.startswith("\\\\"):
            path_to_scan = f"\\\\{path_to_scan}"
        
        self.file_label.config(text=f"Taranıyor: {path_to_scan}", fg="blue")
        self.update_idletasks()
        try:
            if not os.path.isdir(path_to_scan):
                self.file_label.config(text="Ağ yolu/klasör bulunamadı! Yolu kontrol edin veya manuel seçin.", fg="red"); return
            excel_files = [os.path.join(path_to_scan, f) for f in os.listdir(path_to_scan) if f.lower().endswith('.xlsx')]
            if not excel_files:
                self.file_label.config(text="Bu klasörde Excel dosyası bulunamadı. Manuel bir dosya oluşturun/seçin.", fg="orange"); return
            newest_file = max(excel_files, key=os.path.getmtime)
            self.set_active_file(newest_file)
        except Exception:
             self.file_label.config(text=f"Ağ hatası! Yolu kontrol edin veya dosyayı manuel seçin.", fg="red")

    def set_delete_mode(self, event=None):
        self.current_mode = "SİLME"; self.mode_label.config(text="MOD: SİLME (1)", fg="red"); self.status_label.config(text="Silme modu aktif.", fg="red"); self.barcode_entry.delete(0, tk.END)
    def set_add_mode(self, event=None):
        self.current_mode = "EKLEME"; self.mode_label.config(text="MOD: EKLEME (2)", fg="green"); self.status_label.config(text="Okutma bekleniyor.", fg="navy"); self.barcode_entry.delete(0, tk.END)

    def select_file(self):
        initial_dir = self.network_path_var.get().strip()
        if not os.path.isdir(initial_dir): initial_dir = os.path.expanduser("~")
        filepath = filedialog.asksaveasfilename(initialdir=initial_dir, title="Kaydedilecek Excel Dosyasını Seçin veya Yeni Oluşturun", defaultextension=".xlsx", filetypes=[("Excel Dosyaları", "*.xlsx")])
        if filepath: self.set_active_file(filepath)

    def process_barcode(self, event=None):
        if not self.excel_filepath: messagebox.showwarning("Uyarı", "Lütfen önce bir Excel dosyası seçin veya oluşturun."); return
        barcode = self.barcode_entry.get().strip()
        if not barcode: self.status_label.config(text="Barkod alanı boş bırakılamaz.", fg="red"); return
        if self.current_mode == "EKLEME":
            result = save_to_excel(extract_and_slice_barcode(barcode), self.excel_filepath)
            if result == "eklendi": self.status_label.config(text=f"BAŞARILI!\n'{barcode}' dosyaya eklendi.", fg="green")
            else: self.status_label.config(text=f"HATA!\nDosyaya kaydedilemedi.", fg="red")
        elif self.current_mode == "SİLME":
            result = delete_from_excel(barcode, self.excel_filepath)
            if result == "silindi": self.status_label.config(text=f"BAŞARILI!\n'{barcode}' dosyadan silindi.", fg="green")
            elif result == "bulunamadi": self.status_label.config(text=f"UYARI!\n'{barcode}' dosyada bulunamadı.", fg="orange")
            else: self.status_label.config(text=f"HATA!\nSilme işlemi yapılamadı.", fg="red")
        self.barcode_entry.delete(0, tk.END)

if __name__ == "__main__":
    app = BarcodeApp()
    app.mainloop()